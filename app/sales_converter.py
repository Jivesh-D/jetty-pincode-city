"""Convert wide-format Sales sheet from XLSX to normalized CSV."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook

SALES_SHEET = "Sales"
ID_COLS = ("FSN", "TITLE", "CITY")
METRIC_GMV = "GMV"
METRIC_UNITS = "UNITS"

OUTPUT_COLUMNS = (
    "order_date_time",
    "city",
    "product_id",
    "analytic_business_unit",
    "analytic_super_category",
    "analytic_vertical",
    "brand_csv",
    "units",
    "mrp",
)

DATE_SLASH_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
EXCEL_EPOCH = date(1899, 12, 30)
GRAND_TOTAL_LABEL = "GRAND TOTAL"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return _cell_str(value).upper()


def _parse_order_date(value: Any) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        days = int(value)
        parsed = EXCEL_EPOCH + timedelta(days=days)
        return parsed.strftime("%Y-%m-%d")

    text = _cell_str(value)
    match = DATE_SLASH_RE.match(text)
    if match:
        day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
        try:
            return date(year, month, day).strftime("%Y-%m-%d")
        except ValueError:
            return None

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


def _is_grand_total_row(fsn: str, title: str, city: str) -> bool:
    return any(
        _cell_str(value).upper() == GRAND_TOTAL_LABEL for value in (fsn, title, city)
    )


def _metric_number(value: Any) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "0"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    text = _cell_str(value)
    if not text:
        return "0"
    try:
        num = float(text.replace(",", ""))
        if num.is_integer():
            return str(int(num))
        return str(num)
    except ValueError:
        return "0"


def _find_id_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    indices: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        label = _normalize_header(cell)
        if label in ID_COLS and label not in indices:
            indices[label] = idx
    missing = [col for col in ID_COLS if col not in indices]
    if missing:
        raise ValueError(
            f"Sales sheet row 2 must include columns: {', '.join(ID_COLS)}. "
            f"Missing: {', '.join(missing)}"
        )
    return indices


def _build_date_metric_pairs(
    date_row: tuple[Any, ...],
    header_row: tuple[Any, ...],
) -> list[tuple[str, int, int]]:
    pairs: list[tuple[str, int, int]] = []
    pending_date: str | None = None
    pending_gmv: int | None = None

    max_len = max(len(date_row), len(header_row))
    for col in range(max_len):
        header = _normalize_header(header_row[col] if col < len(header_row) else None)
        if header == METRIC_GMV:
            date_val = date_row[col] if col < len(date_row) else None
            parsed = _parse_order_date(date_val)
            if parsed:
                pending_date = parsed
            pending_gmv = col
            continue

        if header == METRIC_UNITS:
            if pending_gmv is None or pending_date is None:
                raise ValueError(
                    "Sales sheet row 2 must have GMV immediately before each Units column"
                )
            pairs.append((pending_date, pending_gmv, col))
            pending_gmv = None
            continue

    if not pairs:
        raise ValueError(
            "No GMV/Units column pairs found on Sales sheet row 2 after FSN, Title, City"
        )

    return pairs


def convert_sales_xlsx_to_csv(file_bytes: bytes) -> str:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if SALES_SHEET not in workbook.sheetnames:
            raise ValueError(f"Workbook must contain a sheet named '{SALES_SHEET}'")

        worksheet = workbook[SALES_SHEET]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if len(rows) < 3:
        raise ValueError("Sales sheet must have a date row, header row, and at least one data row")

    date_row = rows[0]
    header_row = rows[1]
    id_cols = _find_id_columns(header_row)
    metric_pairs = _build_date_metric_pairs(date_row, header_row)

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(OUTPUT_COLUMNS)

    for data_row in rows[2:]:
        fsn = _cell_str(data_row[id_cols["FSN"]] if id_cols["FSN"] < len(data_row) else None)
        title = _cell_str(
            data_row[id_cols["TITLE"]] if id_cols["TITLE"] < len(data_row) else None
        )
        city = _cell_str(data_row[id_cols["CITY"]] if id_cols["CITY"] < len(data_row) else None)

        if not fsn and not title and not city:
            continue

        if _is_grand_total_row(fsn, title, city):
            continue

        for order_date, gmv_col, units_col in metric_pairs:
            gmv_val = data_row[gmv_col] if gmv_col < len(data_row) else None
            units_val = data_row[units_col] if units_col < len(data_row) else None

            writer.writerow(
                [
                    order_date,
                    city,
                    fsn,
                    title,
                    "",
                    "",
                    "",
                    _metric_number(units_val),
                    _metric_number(gmv_val),
                ]
            )

    return output.getvalue()
